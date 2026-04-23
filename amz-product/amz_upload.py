"""
Amazon Product Upload Generator

Reads GROUPIDs from groupids.txt, fetches product data from the database
(skusummary + skumap), and injects rows into a copy of the SHOES.xlsm
template to produce AMZ-Upload.xlsm ready for Amazon Seller Central upload.

Usage:
    python amz_upload.py

Input:
    groupids.txt  - One GROUPID per line (e.g. JLH455-CHARL-NAVY)

Output:
    AMZ-Upload.xlsm - Ready to upload via Amazon Seller Central
                      (Catalog > Add Products via Upload, accepts Excel)

Template columns populated (SHOES.xlsm):
    Col A   (0)   SKU
    Col B   (1)   Product Type          = SHOES
    Col C   (2)   Listing Action        = partial_update
    Col H   (7)   Brand Name            from skusummary.brand
    Col I   (8)   Product Id Type       = EAN
    Col J   (9)   Product Id            from skumap.ean (stripped of B suffix)
    Col FL  (167) Item Condition        = New
    Col FN  (169) List Price with Tax   from skusummary.rrp
    Col GK  (192) Fulfillment Channel   = AMAZON_EU
    Col GP  (197) Your Price GBP        = RRP (same as list price)

SKU logic:
    - If skumap.sku is populated, use it
    - Otherwise generate from skumap.code + '-YYMM' (current month)
"""

import os
import sys
import shutil
from datetime import datetime

import openpyxl
import psycopg2
from dotenv import load_dotenv

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, "SHOES.xlsm")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "AMZ-Upload.xlsm")
GROUPIDS_PATH = os.path.join(SCRIPT_DIR, "groupids.txt")

# Column positions in SHOES.xlsm Template sheet (0-based -> 1-based for openpyxl)
COL_SKU = 1            # A
COL_PRODUCT_TYPE = 2   # B
COL_LISTING_ACTION = 3 # C
COL_BRAND = 8          # H
COL_PROD_ID_TYPE = 9   # I
COL_PROD_ID = 10       # J
COL_CONDITION = 168    # FL
COL_LIST_PRICE = 170   # FN
COL_FULFILLMENT = 193  # GK
COL_PRICE = 198        # GP
COL_BATTERIES_REQ = 267  # JG - "Are batteries required?"

# Data rows start at row 7 in the template (rows 1-6 are settings/headers/examples)
DATA_START_ROW = 7


def get_db_connection():
    """Connect to PostgreSQL using .env credentials."""
    load_dotenv(os.path.join(os.path.dirname(SCRIPT_DIR), ".env"))
    return psycopg2.connect(
        host=os.environ["DB_HOST"],
        port=os.environ.get("DB_PORT", "5432"),
        dbname=os.environ["DB_NAME"],
        user=os.environ["DB_USER"],
        password=os.environ["DB_PASSWORD"],
    )


def read_groupids():
    """Read GROUPIDs from groupids.txt, one per line."""
    if not os.path.exists(GROUPIDS_PATH):
        print(f"Error: {GROUPIDS_PATH} not found")
        sys.exit(1)

    with open(GROUPIDS_PATH, "r") as f:
        ids = [line.strip() for line in f if line.strip()]

    if not ids:
        print("Error: groupids.txt is empty")
        sys.exit(1)

    return ids


def fetch_existing_codes(conn, groupids):
    """Return set of codes already in amzfeed for the given GROUPIDs."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT code FROM amzfeed WHERE groupid = ANY(%s)",
            (groupids,),
        )
        return {row[0] for row in cur.fetchall()}


def fetch_product_data(conn, groupids):
    """Fetch brand/rrp from skusummary and variant data from skumap."""
    rows = []
    skipped = 0
    yymm = datetime.now().strftime("%y%m")
    existing_codes = fetch_existing_codes(conn, groupids)

    with conn.cursor() as cur:
        for gid in groupids:
            # Get summary data
            cur.execute(
                "SELECT brand, rrp FROM skusummary WHERE groupid = %s", (gid,)
            )
            summary = cur.fetchone()
            if not summary:
                print(f"Warning: GROUPID '{gid}' not found in skusummary, skipping")
                continue

            brand, rrp = summary
            if not brand:
                print(f"Warning: No brand for GROUPID '{gid}', skipping")
                continue

            try:
                rrp_val = float(rrp) if rrp else 0.0
            except ValueError:
                print(f"Warning: Invalid RRP '{rrp}' for GROUPID '{gid}', skipping")
                continue

            # Get all active variants
            cur.execute(
                """SELECT sku, code, ean
                   FROM skumap
                   WHERE groupid = %s AND deleted = 0
                   ORDER BY code""",
                (gid,),
            )
            variants = cur.fetchall()

            if not variants:
                print(f"Warning: No active variants for GROUPID '{gid}', skipping")
                continue

            for sku, code, ean in variants:
                # Skip variants already on Amazon
                if code in existing_codes:
                    skipped += 1
                    continue

                # SKU: use existing or generate
                if sku and sku.strip():
                    amz_sku = sku.strip()
                else:
                    amz_sku = f"{code}-{yymm}"

                # EAN: strip trailing B
                clean_ean = ean.rstrip("B") if ean else ""
                if not clean_ean:
                    print(f"Warning: No EAN for code '{code}', skipping variant")
                    continue

                rows.append({
                    "sku": amz_sku,
                    "code": code,
                    "brand": brand,
                    "ean": clean_ean,
                    "rrp": rrp_val,
                })

    if skipped:
        print(f"Skipped {skipped} variant(s) already in Amazon (amzfeed)")

    return rows


def update_skumap(conn, rows):
    """Update skumap with SKU, status, and date so PowerBuilder knows it's on Amazon."""
    today = datetime.now().strftime("%Y-%m-%d")
    updated = 0
    with conn.cursor() as cur:
        for row in rows:
            cur.execute(
                """UPDATE skumap
                   SET sku = %s, status = '1', updated = %s
                   WHERE UPPER(code) = UPPER(%s)""",
                (row["sku"], today, row["code"]),
            )
            updated += cur.rowcount
    conn.commit()
    print(f"Updated {updated} row(s) in skumap (sku, status, updated)")


def generate_upload(rows):
    """Copy SHOES.xlsm template and inject product rows."""
    if not rows:
        print("No rows to write. Check groupids.txt and database.")
        sys.exit(1)

    shutil.copy2(TEMPLATE_PATH, OUTPUT_PATH)
    wb = openpyxl.load_workbook(OUTPUT_PATH, keep_vba=True)
    ws = wb["Template"]

    # Clear any existing data rows (from previous runs via the template example row)
    for r in range(DATA_START_ROW, DATA_START_ROW + 1000):
        if ws.cell(row=r, column=COL_SKU).value is None:
            break
        for col in [COL_SKU, COL_PRODUCT_TYPE, COL_LISTING_ACTION, COL_BRAND,
                     COL_PROD_ID_TYPE, COL_PROD_ID, COL_CONDITION,
                     COL_LIST_PRICE, COL_FULFILLMENT, COL_PRICE,
                     COL_BATTERIES_REQ]:
            ws.cell(row=r, column=col, value=None)

    # Write data rows
    for i, row in enumerate(rows):
        r = DATA_START_ROW + i
        ws.cell(row=r, column=COL_SKU, value=row["sku"])
        ws.cell(row=r, column=COL_PRODUCT_TYPE, value="SHOES")
        ws.cell(row=r, column=COL_LISTING_ACTION, value="partial_update")
        ws.cell(row=r, column=COL_BRAND, value=row["brand"])
        ws.cell(row=r, column=COL_PROD_ID_TYPE, value="EAN")
        ws.cell(row=r, column=COL_PROD_ID, value=row["ean"])
        ws.cell(row=r, column=COL_CONDITION, value="New")
        ws.cell(row=r, column=COL_LIST_PRICE, value=row["rrp"])
        ws.cell(row=r, column=COL_FULFILLMENT, value="AMAZON_EU")
        ws.cell(row=r, column=COL_PRICE, value=row["rrp"])
        ws.cell(row=r, column=COL_BATTERIES_REQ, value="No")

    wb.save(OUTPUT_PATH)
    print(f"Created {OUTPUT_PATH}")
    print(f"  {len(rows)} variants across {len(set(r['brand'] for r in rows))} brand(s)")


def main():
    print("Amazon Product Upload Generator")
    print("=" * 40)

    groupids = read_groupids()
    print(f"Read {len(groupids)} GROUPID(s) from groupids.txt")

    conn = get_db_connection()
    try:
        rows = fetch_product_data(conn, groupids)
        print(f"Fetched {len(rows)} variant(s) from database")
        generate_upload(rows)
        update_skumap(conn, rows)
    finally:
        conn.close()

    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    if os.path.isdir(downloads):
        dest = os.path.join(downloads, "AMZ-Upload.xlsm")
        shutil.copy2(OUTPUT_PATH, dest)
        print(f"Copied to {dest}")
    else:
        print(f"Downloads folder not found at {downloads}; skipped copy")

    print("Done. Upload AMZ-Upload.xlsm to Amazon Seller Central.")


if __name__ == "__main__":
    main()
